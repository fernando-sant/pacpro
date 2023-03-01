# Ce code sert à transformer les données brutes provenant de https://transparency.entsoe.eu/generation/r2/actualGenerationPerProductionType/show
# en un fichier plus exploitable par notre code.
# CLEF 14.IV.2022

# IMPORTANT : les fichiers à transformer doivent être renommés "CodePays_Année_entsoe.xlsx"

from xmlrpc.client import DateTime
import pandas as pd
import math
import os
import numpy as np
from datetime import datetime, timedelta
import openpyxl as xl

# Codes des pays

FR = 'France'
BE = 'Belgium'
DE = 'Germany'
IT = 'Italy'
SP = 'Spain'
UK = 'United Kingdom'
UE = 'UE28'


# Définition du périmètre

country = 'FR'
year = 2019


# Suppression des premières lignes pour que pandas comprenne de quoi il s'agit

wb = xl.load_workbook(country + '_' + str(year) + '_entsoe.xlsx')
sheet = wb['Sheet1']
# Si le code a déjà été lancé il n'y a plus besoin de refaire ces modifications
if sheet['A4'].value == None:
    sheet.unmerge_cells('A1:C1')
    sheet.unmerge_cells('A2:C2')
    sheet.unmerge_cells('A3:C3')
    sheet.unmerge_cells('A5:A8')
    sheet.delete_rows(1, 4)
    sheet.delete_rows(2, 3)
    wb.save(country + '_' + str(year) + '_entsoe.xlsx')

# Importation des données
# Copie du fichier (pour préserver le fichier original)
# Suppression de la variable contenant le fichier original (pour libérer la mémoire)

donnee_brute_de_decoffrage = pd.read_excel(
    country + '_' + str(year) + '_entsoe.xlsx')
donnee_brute = donnee_brute_de_decoffrage.copy()
del donnee_brute_de_decoffrage


# Création de la colonne 'date'

n = donnee_brute.shape

last_date = datetime(year, 1, 1, 0, 0)
date = []


for h in range(n[0]):
    premiere_case = donnee_brute.loc[h, 'MTU']
    if pd.isnull(premiere_case):
        donnee_brute.drop(h, axis=0, inplace=True)
    else:
        if premiere_case[6::] == str(year):
            # last_date = premiere_case[6::] + '-' + \
            #     premiere_case[3:5:] + '-' + premiere_case[:2:]
            last_date = datetime(
                int(premiere_case[6::]), int(premiere_case[3:5:]), int(premiere_case[:2:]))
            donnee_brute.drop(h, axis=0, inplace=True)
        else:
            # heure.append(donnee_brute.loc[h, 'MTU'][:5:])
            horaire = donnee_brute.loc[h, 'MTU'][:5:]
            minute = horaire[3::]
            if minute[0] == 0:
                minute = int(minute[1])
            heure = horaire[:2:]
            if heure[0] == 0:
                heure = int(heure[1])
            temps = int(minute) + 60 * int(heure)
            date.append(last_date + timedelta(minutes=temps))

donnee_brute.drop(['MTU'], axis=1, inplace=True)

donnee_brute['date'] = date
# donnee_brute['heure'] = heure


# Gestion des données manquantes

donnee_brute = donnee_brute.replace(['n/e'], 0)


# On s'assure que les données du mix sont bien en format numérique

columns = donnee_brute.columns
columns_numeric = list(columns)
columns_numeric.remove('date')
# columns_numeric.remove('heure')
donnee_brute[columns_numeric] = donnee_brute[columns_numeric].apply(
    pd.to_numeric)


# Gestion des données manquantes

# Gestion du cas où la première case d'une colonne est vide :

for column in columns_numeric:
    if math.isnan(donnee_brute.loc[0, column]):
        donnee_brute.loc[0, column] = 0


# Propagation de la dernière valeur valide

donnee_brute["Hydro Pumped Storage"].fillna(0, inplace=True)
donnee_brute.fillna(method='ffill', inplace=True)


# Fusion des colonnes

donnee_brute['wind'] = donnee_brute['Wind Offshore'] + \
    donnee_brute['Wind Onshore']

donnee_brute['hydro'] = donnee_brute['Hydro Run-of-river and poundage'] + \
    donnee_brute['Hydro Water Reservoir'] + \
    donnee_brute['Marine'] + \
    donnee_brute['Hydro Pumped Storage']

donnee_brute['coal'] = donnee_brute['Fossil Brown coal/Lignite'] + \
    donnee_brute['Fossil Hard coal'] + \
    donnee_brute['Fossil Oil shale'] + \
    donnee_brute['Fossil Peat']

donnee_brute['natural_gas'] = donnee_brute['Fossil Gas'] + \
    donnee_brute['Fossil Coal-derived gas']


# Suppression des colonnes inutiles

for j in range(len(columns)):
    if columns[j][:7] == 'Unnamed':
        donnee_brute.drop(columns[j], axis=1, inplace=True)

donnee_brute.drop('Hydro Pumped Storage', axis=1, inplace=True)
donnee_brute.drop('Wind Offshore', axis=1, inplace=True)
donnee_brute.drop('Wind Onshore', axis=1, inplace=True)
donnee_brute.drop('Hydro Run-of-river and poundage', axis=1, inplace=True)
donnee_brute.drop('Hydro Water Reservoir', axis=1, inplace=True)
donnee_brute.drop('Marine', axis=1, inplace=True)
donnee_brute.drop('Fossil Brown coal/Lignite', axis=1, inplace=True)
donnee_brute.drop('Fossil Coal-derived gas', axis=1, inplace=True)
donnee_brute.drop('Fossil Gas', axis=1, inplace=True)
donnee_brute.drop('Fossil Hard coal', axis=1, inplace=True)
donnee_brute.drop('Fossil Oil shale', axis=1, inplace=True)
donnee_brute.drop('Fossil Peat', axis=1, inplace=True)


# On renomme les colonnes pour être plus clair

donnee_brute.rename(columns={'Biomass': 'biomass'}, inplace=True)
donnee_brute.rename(columns={'Geothermal': 'geothermal'}, inplace=True)
donnee_brute.rename(columns={'Nuclear': 'nuclear'}, inplace=True)
donnee_brute.rename(columns={'Other': 'other'}, inplace=True)
donnee_brute.rename(
    columns={'Other renewable': 'other_renewable'}, inplace=True)
donnee_brute.rename(columns={'Solar': 'solar'}, inplace=True)
donnee_brute.rename(columns={'Waste': 'waste'}, inplace=True)
donnee_brute.rename(columns={'Fossil Oil': 'oil'}, inplace=True)


# Ajout de la colonne 'total'

donnee_brute['total'] = donnee_brute['biomass'] +\
    donnee_brute['natural_gas'] +\
    donnee_brute['coal'] +\
    donnee_brute['oil'] +\
    donnee_brute['geothermal'] +\
    donnee_brute['nuclear'] +\
    donnee_brute['other'] +\
    donnee_brute['other_renewable'] +\
    donnee_brute['solar'] +\
    donnee_brute['waste'] +\
    donnee_brute['wind'] +\
    donnee_brute['hydro']

# Si la ligne ne correspond pas à une heure entière, on la supprime (pour avoir le même format pour tous les pays)

for i in donnee_brute.index:
    if str(donnee_brute.loc[i, 'date'])[14:16:] != '00':
        donnee_brute.drop(i, axis=0, inplace=True)

# Sauvegarde dans un nouveau fichier excel

donnee_brute.to_excel(str(os.getcwd()) + '\\' + country +
                      '_' + str(year) + '_electric_mix.xlsx', index=False)
